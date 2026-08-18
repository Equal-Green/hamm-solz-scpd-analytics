"""Generate the Annex 2 #2 Data Request Tracking Matrix and file inventory.

Produces one workbook seeded with everything derivable from the delivered
archive, so the Consultant only has to fill the columns that require their own
correspondence record (request dates, follow-up actions, Circular EP contact).

    .venv/bin/python tools/build_tracking_matrix.py [-o OUTPUT.xlsx]

Sheets
  Read me            -- what this is, what the Consultant must complete
  Request Matrix     -- the 36 Annex 3 items, coverage read from the archive
  File Inventory     -- every delivered file, mapped to its Annex 3 category
  Correspondence     -- Circular EP references found in delivered file names
  Outstanding        -- items with no material delivered; the Day-90 ask list
"""
import argparse
import json
import mmap
import os
import struct
import sys

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from openpyxl import Workbook                                  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter                   # noqa: E402

import compliance as C                                         # noqa: E402
from config import ZIP_PATH                                    # noqa: E402
from pipeline.archive import load_inventory                     # noqa: E402

IRON, CREAM, CRIMSON = "FF141414", "FFF8F5EF", "FFDC2828"
GREEN, AMBER, RED = "FF2E7D5B", "FFE0A106", "FFDC2828"
COVERAGE_LABEL = {
    "received": "Received",
    "partial": "Partial — incomplete",
    "none": "Not delivered",
}
COVERAGE_COLOR = {"received": GREEN, "partial": AMBER, "none": RED}
_THIN = Side(style="thin", color="FFD8D5CF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def file_paths(zip_path=ZIP_PATH):
    """Every file inside the archive. The ZIP is written in streaming mode, so
    local headers carry no usable sizes or original timestamps -- names only."""
    if not os.path.exists(zip_path):
        return []
    out, seen = [], set()
    with open(zip_path, "rb") as f:
        mm = mmap.mmap(f.fileno(), 0, access=mmap.ACCESS_READ)
        try:
            off = 0
            while True:
                p = mm.find(b"PK\x03\x04", off)
                if p == -1:
                    break
                fnl = struct.unpack_from("<H", mm, p + 26)[0]
                name = mm[p + 30:p + 30 + fnl].decode("utf-8", "replace")
                if name.startswith("INFORMACIÓN/") and not name.endswith("/") \
                        and name not in seen:
                    seen.add(name)
                    out.append(name)
                off = p + 4
        finally:
            mm.close()
    return sorted(out)


def _folder_to_categories():
    """Archive folder number -> the Annex 3 categories it feeds."""
    m = {}
    for _ref, cat, _item, folders, _cov, _note in C.ANNEX3_ITEMS:
        for n in folders:
            m.setdefault(n, set()).add(cat)
    return {n: ", ".join(sorted(v)) for n, v in m.items()}


def _header(ws, labels, widths):
    ws.append(labels)
    for i, (label, width) in enumerate(zip(labels, widths), start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True, color=CREAM, size=10)
        cell.fill = PatternFill("solid", fgColor=IRON)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _body(ws, wrap_cols=()):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=cell.column_letter in wrap_cols)
            if cell.font.size is None:
                cell.font = Font(size=10)


def sheet_readme(wb, inv, summary):
    ws = wb.create_sheet("Read me")
    ws.column_dimensions["A"].width = 110
    lines = [
        ("Data Request Tracking Matrix — Ecuador Waste EIPP Project", True),
        ("Annex 2 #2 deliverable · Consultancy Service Agreement, "
         "THE HAMM SOLZ Co., Ltd. ↔ Mr. Carlos Arcos Pastor", False),
        (f"Effective Date {C.EFFECTIVE_DATE:%d %B %Y} · "
         f"{C.TERM_DAYS}-day term ending {C.END_DATE:%d %B %Y}", False),
        ("", False),
        ("WHAT IS ALREADY FILLED IN", True),
        ("Every column that can be established from the delivered archive: the "
         "36 Annex 3 items, which archive folders answer each one, the coverage "
         "reading, the full file inventory, and the Circular EP correspondence "
         "references embedded in delivered file names.", False),
        ("", False),
        ("WHAT THE CONSULTANT MUST COMPLETE", True),
        ("On 'Request Matrix', the four columns marked TO COMPLETE: the Circular "
         "EP contact or department, the date each request was sent, the "
         "follow-up actions taken with dates, and the response status. Those "
         "come from the Consultant's own correspondence record and cannot be "
         "reconstructed from the files.", False),
        ("Also confirm or correct the coverage reading in column E. It is a "
         "first-pass reading of folder contents, not a verified status.", False),
        ("", False),
        ("WHY THE FOLLOW-UP COLUMNS MATTER", True),
        ("Annex 1 §5: where Circular EP does not possess, collect, or disclose "
         "material within the term, that is not non-performance by the "
         "Consultant PROVIDED reasonable efforts and follow-up are documented. "
         "Those columns are the documentation that provision depends on.", False),
        ("", False),
        ("ARCHIVE AS DELIVERED", True),
        (f"{inv.get('total_files', 0):,} files across "
         f"{len(inv.get('folders', []))} topic folders.", False),
        (f"Annex 3 items: {summary['received']} received, "
         f"{summary['partial']} partial, {summary['none']} not delivered "
         f"(of {summary['total']}).", False),
        ("The source ZIP is written in streaming mode, so it carries no usable "
         "original file dates or sizes — the inventory lists names and formats "
         "only.", False),
    ]
    for text, bold in lines:
        ws.append([text])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(bold=bold, size=12 if bold and ws.max_row == 1 else 10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def sheet_request_matrix(wb, inv):
    ws = wb.create_sheet("Request Matrix")
    _header(ws, [
        "Ref", "Annex 3 cat.", "Requested item", "Archive folders answering it",
        "Coverage (confirm)", "Files", "Circular EP contact / dept "
        "[TO COMPLETE]", "Request date [TO COMPLETE]",
        "Follow-up actions + dates [TO COMPLETE]",
        "Response status [TO COMPLETE]", "Notes from archive reading",
    ], [7, 11, 46, 34, 19, 7, 24, 16, 34, 20, 46])

    for r in C.annex3_item_rows(inv):
        ws.append([r["ref"], r["category"], r["item"],
                   "; ".join(r["folders"]) or "—",
                   COVERAGE_LABEL[r["coverage"]],
                   r["files"] or "", "", "", "", "", r["note"]])
        ws.cell(row=ws.max_row, column=5).font = Font(
            bold=True, size=10, color=COVERAGE_COLOR[r["coverage"]])
    _body(ws, wrap_cols=("C", "D", "I", "K"))
    return ws


def sheet_file_inventory(wb, paths):
    ws = wb.create_sheet("File Inventory")
    _header(ws, ["#", "Top-level folder", "Annex 3 cat.", "Sub-path",
                 "File name", "Format", "Status", "Comments"],
            [6, 34, 11, 52, 58, 9, 12, 30])
    cats = _folder_to_categories()
    for i, path in enumerate(paths, start=1):
        parts = path.split("/")
        folder = parts[1] if len(parts) > 2 else "(root)"
        num = C._folder_number(folder)
        sub = "/".join(parts[2:-1])
        ext = os.path.splitext(path)[1].lower().lstrip(".") or "—"
        ws.append([i, folder, cats.get(num, ""), sub, parts[-1], ext,
                   "Received", ""])
    _body(ws, wrap_cols=("D", "E", "H"))
    return ws


def sheet_correspondence(wb):
    ws = wb.create_sheet("Correspondence")
    _header(ws, ["Reference found in delivered file name", "Archive folder",
                 "Subject", "Belongs to this engagement? [TO COMPLETE]",
                 "Notes [TO COMPLETE]"], [38, 34, 56, 30, 40])
    for ref, folder, subject in C.CORRESPONDENCE_REFS:
        ws.append([ref, folder, subject, "", ""])
    _body(ws, wrap_cols=("C", "E"))
    return ws


def sheet_outstanding(wb, inv):
    ws = wb.create_sheet("Outstanding")
    _header(ws, ["Ref", "Annex 3 cat.", "Item not delivered",
                 "Why it matters / note", "Include in final written request?"],
            [7, 11, 56, 56, 30])
    for r in C.annex3_item_rows(inv):
        if r["coverage"] == "none":
            ws.append([r["ref"], r["category"], r["item"], r["note"], "Yes"])
    _body(ws, wrap_cols=("C", "D"))
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("-o", "--output",
                    default=os.path.join(_ROOT, "deliverables",
                                         "Annex2-Data-Request-Tracking-Matrix"
                                         ".xlsx"))
    args = ap.parse_args()

    inv = load_inventory() or {}
    summary = C.annex3_item_summary(inv)
    paths = file_paths()
    if not paths:
        print(f"warning: source ZIP not found at {ZIP_PATH} — "
              "File Inventory sheet will be empty")

    wb = Workbook()
    wb.remove(wb.active)
    sheet_readme(wb, inv, summary)
    sheet_request_matrix(wb, inv)
    sheet_file_inventory(wb, paths)
    sheet_correspondence(wb)
    sheet_outstanding(wb, inv)

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    wb.save(args.output)
    print(f"wrote {args.output}")
    print(f"  Request Matrix   {summary['total']} items "
          f"({summary['received']} received / {summary['partial']} partial / "
          f"{summary['none']} not delivered)")
    print(f"  File Inventory   {len(paths):,} files")
    print(f"  Correspondence   {len(C.CORRESPONDENCE_REFS)} references")
    print(f"  Outstanding      {summary['none']} items to request in writing")


if __name__ == "__main__":
    main()
